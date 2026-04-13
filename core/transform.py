"""
core/transform.py
Transforme ProchainesRencontres.xlsx en MatchsSemaine_YYYYMMDD.xlsx
et retourne les données structurées pour la génération d'affiche.

Sources de référence (dans config/) :
  Divisions.xlsx   — DIVISION → CATEGORIE + TYPEMATCH (exact match)
  NomsEquipes.xlsx — NOMLONG  → NOMCOURT  (fallback : nom brut nettoyé)

Convention placeholders affiche :
  Domicile  → DOM_N_EQUIPE1 (ABP), DOM_N_EQUIPE2 (adv), DOM_N_DATE
  Extérieur → EXT_N_EQUIPE1 (adv), EXT_N_EQUIPE2 (ABP), EXT_N_DATE
  Format date : SAMEDI\n14H00
"""

import re
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CONFIG_DIR  = Path("config")
TEAMS_PATH  = CONFIG_DIR / "teams.json"
DIV_PATH    = CONFIG_DIR / "Divisions.xlsx"
NOMS_PATH   = CONFIG_DIR / "NomsEquipes.xlsx"

_MOIS_FR = {
    1:"janvier", 2:"février", 3:"mars", 4:"avril",
    5:"mai", 6:"juin", 7:"juillet", 8:"août",
    9:"septembre", 10:"octobre", 11:"novembre", 12:"décembre",
}
_JOURS_FR = ["LUNDI","MARDI","MERCREDI","JEUDI","VENDREDI","SAMEDI","DIMANCHE"]


# ──────────────────────────────────────────────
# Chargement des référentiels
# ──────────────────────────────────────────────

def _load_divisions() -> dict:
    """Retourne {division_str: {"categorie": ..., "typematch": ...}}."""
    df = pd.read_excel(DIV_PATH, header=0)
    df.columns = [c.strip().upper() for c in df.columns]
    result = {}
    for _, row in df.iterrows():
        div = str(row["DIVISION"]).strip()
        result[div] = {
            "categorie": str(row["CATEGORIE"]).strip(),
            "typematch": str(row["TYPEMATCH"]).strip(),
        }
    return result


def _load_noms() -> dict:
    """Retourne {nomlong_upper: nomcourt}."""
    df = pd.read_excel(NOMS_PATH, header=0)
    df.columns = [c.strip().upper() for c in df.columns]
    result = {}
    for _, row in df.iterrows():
        long_ = str(row["NOMLONG"]).strip()
        court = str(row["NOMCOURT"]).strip()
        if long_:
            result[long_.upper()] = court
    return result


def _load_abp_marker() -> str:
    try:
        with open(TEAMS_PATH, encoding="utf-8") as f:
            return json.load(f).get("abp_marker", "AMICALE BASKET PECQUENCOURT")
    except Exception:
        return "AMICALE BASKET PECQUENCOURT"


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _format_date_fr(dt) -> str:
    return f"{dt.day} {_MOIS_FR[dt.month]} {dt.year}"


def _format_date_affiche(date_str: str, heure_str: str) -> str:
    """'SAMEDI\n14H00'"""
    try:
        dt = pd.to_datetime(f"{date_str} {heure_str}", dayfirst=True)
        jour   = _JOURS_FR[dt.weekday()]
        heure  = dt.strftime("%H").lstrip("0") or "0"
        minute = dt.strftime("%M")
        return f"{jour}\n{heure}H{minute}"
    except Exception:
        return f"{date_str} {heure_str}"


def _resolve_division(division: str, div_map: dict) -> dict:
    """Correspondance exacte dans Divisions.xlsx.
    Fallback : regex sur le préfixe si pas trouvé (robustesse).
    """
    division = str(division).strip()

    # Correspondance exacte
    if division in div_map:
        return div_map[division]

    # Fallback regex préfixe (ex: "DFU11-8-P2-P2-NOUVEAU" non encore référencé)
    m = re.match(r"^([A-Z]+\d*)", division)
    prefix = m.group(1) if m else division
    for key, val in div_map.items():
        if key.startswith(prefix):
            return val

    return {"categorie": division, "typematch": "CHAMPIONNAT"}


def _is_abp(name: str, marker: str) -> bool:
    return marker.upper() in str(name).upper()


def _abp_short_name(division: str, team_name: str, div_map: dict) -> str:
    """Ex: 'AMICALE BASKET PECQUENCOURT - 2 (10)' + 'DFU11-8-P2-P2' → 'U11F-2'"""
    info     = _resolve_division(division, div_map)
    category = info["categorie"]
    m = re.search(r"PECQUENCOURT\s*-\s*(\d+)", str(team_name), re.IGNORECASE)
    suffix = f"-{m.group(1)}" if m else ""
    return category + suffix


def _opponent_short_name(name: str, noms_map: dict) -> str:
    """Cherche le nom court dans NomsEquipes.xlsx.
    Conserve le suffixe d'équipe adversaire ('- 2').
    Fallback : nom brut nettoyé (tronqué à 25 car.).
    """
    clean = re.sub(r"\s*\(\d+\)\s*$", "", str(name)).strip()
    suffix_match = re.search(r"\s*-\s*(\d+)\s*$", clean)
    team_suffix  = f" - {suffix_match.group(1)}" if suffix_match else ""
    base_name    = clean[: suffix_match.start()].strip() if suffix_match else clean

    # Recherche exacte (insensible à la casse)
    court = noms_map.get(base_name.upper())
    if court:
        return court + team_suffix

    # Recherche partielle — nom le plus long qui correspond
    best = None
    for long_, short_ in sorted(noms_map.items(), key=lambda x: -len(x[0])):
        if long_ in base_name.upper():
            best = short_
            break
    if best:
        return best + team_suffix

    return (base_name + team_suffix)[:25]


def _parse_date_heure(date_val, heure_val) -> tuple:
    date_str  = str(date_val).strip()
    heure_str = str(heure_val).strip()
    try:
        dt = pd.to_datetime(f"{date_str} {heure_str}", dayfirst=True)
        return dt.strftime("%Y%m%d%H%M"), dt.strftime("%d/%m/%Y"), dt.strftime("%H:%M")
    except Exception:
        return f"{date_str}{heure_str}", date_str, heure_str


# ──────────────────────────────────────────────
# API publique
# ──────────────────────────────────────────────

def load_planning_data(source_path: str):
    """Lit ProchainesRencontres.xlsx.

    Retourne (domicile_list, exterieur_list, week_label).

    Chaque match :
        equipe, adversaire, typematch,
        date_affiche ('SAMEDI\\n14H00'), date, heure, salle, domicile
    """
    div_map  = _load_divisions()
    noms_map = _load_noms()
    marker   = _load_abp_marker()

    df = pd.read_excel(source_path, header=0)
    df.columns = [
        "Division","NumMatch","Equipe1","Equipe2",
        "Date","Heure","Salle","Emarque",
        "Score1","Forfait1","Score2","Forfait2",
    ]

    domicile_list  = []
    exterieur_list = []
    dates_found    = []

    for _, row in df.iterrows():
        e1, e2 = str(row["Equipe1"]), str(row["Equipe2"])
        if "exempt" in e1.lower() or "exempt" in e2.lower():
            continue

        division = str(row["Division"]).strip()
        info     = _resolve_division(division, div_map)
        typematch = info["typematch"]

        sort_key, date_str, heure_str = _parse_date_heure(row["Date"], row["Heure"])
        salle        = str(row["Salle"]) if pd.notna(row["Salle"]) else ""
        date_affiche = _format_date_affiche(date_str, heure_str)

        if _is_abp(e1, marker):
            abp      = _abp_short_name(division, e1, div_map)
            opp      = _opponent_short_name(e2, noms_map)
            domicile = True
        elif _is_abp(e2, marker):
            abp      = _abp_short_name(division, e2, div_map)
            opp      = _opponent_short_name(e1, noms_map)
            domicile = False
        else:
            continue

        try:
            dates_found.append(pd.to_datetime(date_str, dayfirst=True))
        except Exception:
            pass

        match = {
            "equipe":       abp,
            "adversaire":   opp,
            "typematch":    typematch,
            "date_affiche": date_affiche,
            "date":         date_str,
            "heure":        heure_str,
            "salle":        salle,
            "domicile":     domicile,
            "_sort":        sort_key,
        }
        (domicile_list if domicile else exterieur_list).append(match)

    domicile_list.sort(key=lambda m: m["_sort"])
    exterieur_list.sort(key=lambda m: m["_sort"])

    if dates_found:
        first      = min(dates_found)
        week_label = f"Semaine du {_format_date_fr(first)}"
    else:
        week_label = f"Semaine du {_format_date_fr(datetime.now())}"

    return domicile_list, exterieur_list, week_label


def transform(source_path: str, output_path: str = "") -> str:
    """Génère MatchsSemaine_YYYYMMDD.xlsx. Retourne le chemin du fichier créé."""
    domicile_list, exterieur_list, _ = load_planning_data(source_path)
    all_matches = sorted(domicile_list + exterieur_list, key=lambda m: m["_sort"])

    if not output_path:
        date_tag    = datetime.now().strftime("%Y%m%d")
        output_path = f"MatchsSemaine_{date_tag}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers    = ["Type","Equipe1","Equipe2","Date","Salle",
                  "Arbitres","Chrono","Emarque","Souffleur","RespSalle","Buvette"]
    col_widths = [13, 14, 14, 18, 24, 14, 10, 12, 12, 12, 10]

    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, hdr_align
    ws.row_dimensions[1].height = 22

    alt_fill  = PatternFill("solid", start_color="DCE6F1")
    std_font  = Font(name="Arial", size=10)
    std_align = Alignment(vertical="center")

    for r_idx, m in enumerate(all_matches, 2):
        eq1 = m["equipe"]    if m["domicile"] else m["adversaire"]
        eq2 = m["adversaire"] if m["domicile"] else m["equipe"]
        row_data = [m["typematch"], eq1, eq2,
                    f"{m['date']} {m['heure']}", m["salle"],
                    "", "", "", "", "", ""]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.font, c.alignment = std_font, std_align
            if r_idx % 2 == 0:
                c.fill = alt_fill
        ws.row_dimensions[r_idx].height = 18

    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    wb.save(output_path)
    return output_path
